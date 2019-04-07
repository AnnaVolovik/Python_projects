import itertools
import os
import re
import time

from bs4 import BeautifulSoup
from flask import render_template, request
import openpyxl
from openpyxl.styles import Font, Border, Side
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait

from app import app, db
from app.models import DentistContacts


def xpath_soup(element):
    """
    Generate xpath of a beautiful soup element
    :param element: bs4 text or node
    :return: xpath as string
    """
    components = []
    child = element if element.name else element.parent
    for parent in child.parents:
        """
        @type parent: bs4.element.Tag
        """
        previous = itertools.islice(parent.children, 0, parent.contents.index(child))
        xpath_tag = child.name
        xpath_index = sum(1 for i in previous if i.name == xpath_tag) + 1
        components.append(xpath_tag if xpath_index == 1 else '%s[%d]' % (xpath_tag, xpath_index))
        child = parent
    components.reverse()
    return '/%s' % '/'.join(components)

def parse_block(contact_section):
    """ Parse a Beautiful Soup element, return data as a dictionary
    :param contact_section - Beautiful Soup object
    :return dict()
    """
    ancor_div = contact_section.find('div', {'class': 'Place__mainTitle'})
    ancor = ancor_div.find('a')

    website = ancor.get('href')
    name = ancor.text
    address = contact_section.find('span', {'class': 'Place__addressText'}).text
    phone_div = contact_section.find('div', {'class': 'Place__phoneNumber '})
    try:
        ancor_phone = phone_div.find('a')
        phone = ancor_phone.get('href').replace('tel:', '')
    except:
        phone = None

    return dict(
        name=name,
        address=address,
        phone=phone,
        website=website
)

def export_results_to_excel(res):

    """Export results to an excel file, save it in static folder
    :param res: dict list
    :return 'OK', 200
    """

    # set output file and its styles
    wb = openpyxl.Workbook()
    ws = wb.active  # grab the active worksheet

    # add and style the header, set column width
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 25

    ws['A1'].value = 'Company name'
    ws['B1'].value = 'Address'
    ws['C1'].value = 'Website'
    ws['D1'].value = 'Phone number'

    bold = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = bold

    row = 2
    for tmp in res:
        ws.cell(row=row, column=1).value = tmp['name']
        ws.cell(row=row, column=2).value = tmp['address']
        ws.cell(row=row, column=3).value = tmp['website']
        ws.cell(row=row, column=4).value = tmp['phone']
        row += 1

    # add total row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = 'Total: '
    ws.cell(row=row, column=4).value = f'{len(res)} entries'

    # # add borders
    for row in range(1, row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = Border(
                top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))

    # save file to the app's static folder
    file_name = os.path.join('static', 'project_two.xlsx')
    wb.save(os.path.join(app.root_path, file_name))
    wb.close()

    return file_name

def export_results_to_db(res):
    """
    Save the results to sqllite database
    :param res: dict list
    :return: "OK", 200
    """
    for tmp in res:

        # if we have this clinic - update its info
        q = db.session.query(DentistContacts).filter(
                DentistContacts.name == tmp['name'])
        if q.count() > 1:
            q = q.filter(DentistContacts.address == tmp['address'])
        if q.count() == 1:
            entry = q.one()
            setattr(entry, 'address', tmp['address'])
            setattr(entry, 'website', tmp['website'])
            if tmp.get('phone', None):
                setattr(entry, 'phone', tmp['phone'])
        else:  # else - add new entry
            entry = DentistContacts(**tmp)
            db.session.add(entry)
    try:
        db.session.flush()
        db.session.commit()
    except:
        db.session.rollback()
        return render_template('project_two.html', error='Data Base Error. Please try init_db')


@app.route('/project_two', methods=['GET', 'POST'])
def project_two():

    """ Iterate over pages and scrape the information needed """
    if request.method == 'GET':
        return render_template('project_two.html')

    url = "https://www.103.ua/list/stomatologii/kiev/"
    page_limit = 3  # by default for demonstration purposes we only go through first two pages

    res = []

    driver = None

    # attempt to parse with selenium
    try:
        driver = webdriver.Firefox(executable_path='/usr/local/bin/geckodriver')
        driver.get(url)

        pages = driver.find_elements_by_class_name('Pagination__page')

        for i in range(1, page_limit):
            pages[i].click()
            time.sleep(3)

            html = driver.page_source
            soup = BeautifulSoup(html)

            # find all contact information blocks
            for contact_block in soup.find_all('div', {'class': 'Place__wrap'}):
                # try to parse without opening the pop up window
                ancor_div = contact_block.find('div', {'class': 'Place__mainTitle'})
                ancor = ancor_div.find('a')

                website = ancor.get('href')
                name = ancor.text
                address = contact_block.find('span', {'class': 'Place__addressText'}).text
                phone_div = contact_block.find('div', {'class': 'Place__phoneNumber '})
                try:
                    ancor_phone = phone_div.find('a')
                    phone = ancor_phone.get('href').replace('tel:', '')
                except:
                    # open the pop up
                    # phone = None
                    contacts_button = contact_block.find('button', {'class': 'Place__showContacts'})
                    xpath = xpath_soup(contacts_button)
                    selenium_button = driver.find_element_by_xpath(xpath)
                    selenium_button.click()

                    # assume new code will appear in the same window, wait explicitly to locate elements
                    WebDriverWait(driver, 15).until(
                        lambda x: x.find_element_by_class_name('PopupWrapper'))

                    # phone number
                    try:
                        phone_list = driver.find_elements_by_class_name('ContactsPopupPhones__number')
                        phone_links = [x.get_attribute('href') for x in phone_list]
                        phone = ', '.join([re.findall('tel:(.+)', x)[0] for x in phone_links])
                    except:
                        phone = driver.find_element_by_class_name('PhoneLink__number').get_attribute('innerHTML')

                    # close the pop up
                    webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                    time.sleep(2)

                res.append(dict(
                    name=name,
                    address=address,
                    phone=phone,
                    website=website
                    ))

        driver.close()
        driver.quit()

    # backup part written on requests & beautiful soup if selenium fails
    except:
        if driver:
            # close the driver if it was open
            driver.close()
            driver.quit()

        # get the same information with requests
        for page_number in range(page_limit):
            page_url = f'{url}?page={page_number}'

            r = requests.get(page_url)
            soup = BeautifulSoup(r.content, 'html.parser')

            # find all contact information blocks
            for contact_block in soup.find_all('div', {'class': 'Place__wrap'}):
                res.append(parse_block(contact_block))

    # sort results by company name
    res.sort(key=lambda x: x['name'])

    # export results to database
    export_results_to_db(res)

    # writing elements into MS Excel file
    excel_file = export_results_to_excel(res)

    return render_template('project_two.html', entries=res, file_name=excel_file)

